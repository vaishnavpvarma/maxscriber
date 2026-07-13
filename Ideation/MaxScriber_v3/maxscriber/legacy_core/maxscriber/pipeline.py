"""
MaXScriber v1.0
Pipeline orchestration: logging, extraction, aggregation, and reporting.
Supports both full pipeline and individual subcommand execution.
"""

import concurrent.futures
import logging
import os
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List

import pandas as pd
from alive_progress import alive_bar
from openpyxl.styles import Font, PatternFill

from maxscriber.constants import OUTPUT_COLUMNS

# =============================================================================
# INTERMEDIATE DATA PERSISTENCE
# =============================================================================
from maxscriber.db import MaxScriberDB
from maxscriber.extraction import (
    determine_tests_done,
    extract_dates_from_content,
    extract_metadata,
    extract_pdf_content,
    func_dengue_dedicated,
    func_line_pattern,
    func_table_smart,
)
from maxscriber.plots import generate_clinical_plots
from maxscriber.stats import generate_condensed_stats
from maxscriber.stratification import generate_pzfx, perform_stratification, run_kruskal_wallis


def get_db(output_dir: Path, job_name: str) -> MaxScriberDB:
    return MaxScriberDB(output_dir / f"{job_name}_database.sqlite")


def save_extraction_data(output_dir: Path, payload: dict, job_name: str):
    """Save intermediate extraction data for subcommand independence."""
    db = get_db(output_dir, job_name)
    db.save_pipeline_state("failed_files", payload["failed_files"])
    db.save_pipeline_state("qc_duplicates", payload["qc_duplicates"])
    db.save_pipeline_state("longitudinal_tests", payload.get("longitudinal_tests", {}))

    # We also save the aggregated data as pipeline state for fast loading without complex joins
    db.save_pipeline_state("patient_data", payload["patient_data"])
    db.save_pipeline_state("latest_data", payload["latest_data"])
    # Convert sets to lists for JSON serialization
    f_map = {k: list(v) for k, v in payload["f_map"].items()}
    db.save_pipeline_state("f_map", f_map)
    db.save_pipeline_state("m_map", payload["m_map"])

    # Insert individual extracted data into relational tables in a single transaction
    db.insert_extractions_bulk(payload["all_extractions"])


def load_extraction_data(output_dir: Path, job_name: str) -> dict:
    """Load previously saved extraction data."""
    db_path = output_dir / f"{job_name}_database.sqlite"
    if not db_path.exists():
        print(f"ERROR: No database found at {db_path}")
        print("Run 'maxscriber transcribe' first to generate extraction data.")
        sys.exit(1)

    db = get_db(output_dir, job_name)

    payload = {
        "all_extractions": db.get_all_extractions(),
        "failed_files": db.load_pipeline_state("failed_files", []),
        "qc_duplicates": db.load_pipeline_state("qc_duplicates", []),
        "longitudinal_tests": db.load_pipeline_state("longitudinal_tests", {}),
        "patient_data": db.load_pipeline_state("patient_data", {}),
        "latest_data": db.load_pipeline_state("latest_data", {}),
        "m_map": db.load_pipeline_state("m_map", {}),
    }
    f_map_json = db.load_pipeline_state("f_map", {})
    payload["f_map"] = {k: set(v) for k, v in f_map_json.items()}

    return payload


# =============================================================================
# LOGGING
# =============================================================================


def setup_logging(
    output_dir: Path, job_name: str, append: bool = False, verbose: bool = False
) -> logging.Logger:
    """Setup comprehensive logging (V2 style).

    Args:
        output_dir: Directory for the log file.
        job_name: Prefix for log file.
        append: If True, append to existing log (for Phase 2).
                If False, start fresh (for Phase 1).
        verbose: If True, log to stdout as well.
    """
    log_file = output_dir / f"{job_name}_extraction.log"

    # Clear any existing handlers to avoid duplicate output on re-runs
    root = logging.getLogger()
    for h in root.handlers[:]:
        root.removeHandler(h)

    file_mode = "a" if append else "w"
    handlers = [logging.FileHandler(log_file, mode=file_mode)]
    if verbose:
        handlers.append(logging.StreamHandler(sys.stdout))

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=handlers,
    )
    logger = logging.getLogger("MaxScriber")

    if append:
        logger.info("")
        logger.info("=" * 80)
        logger.info("Resuming pipeline — Phase 2 (QC / Stats / Plot)")
        logger.info("=" * 80)
    else:
        # Banner in log
        logger.info("=" * 80)
        logger.info("Intelligent Multi-Pass Medical PDF Extractor")
        logger.info("No AI/ML - Pure Rule-Based with Cross-Validation")
        logger.info("=" * 80)
    return logger


# =============================================================================
# EXTRACTION PHASE
# =============================================================================


def _process_pdf_worker(pdf: Path, output_dir: Path, job_name: str, verbose: bool):
    """
    Worker function for processing a single PDF. Extracted to module level
    so it can be pickled by multiprocessing (ProcessPoolExecutor).
    """
    local_logger = setup_logging(output_dir, job_name, append=True, verbose=verbose)
    local_logger.info(f"\n{'=' * 80}\nProcessing: {pdf.name}\n{'=' * 80}")

    content = extract_pdf_content(str(pdf), local_logger)
    if not content:
        return pdf.name, None, None, None

    meta = extract_metadata(content, local_logger)
    dates = extract_dates_from_content(content, local_logger)

    res_table = func_table_smart(content, dates, local_logger)
    res_line = func_line_pattern(content, dates, local_logger)
    res_dengue = func_dengue_dedicated(content, dates, local_logger)

    votes = defaultdict(lambda: defaultdict(list))
    strategies = [("Table", res_table), ("Line", res_line), ("Dengue-Dedicated", res_dengue)]
    for s_name, res in strategies:
        for t, dv in res.items():
            for d, v in dv.items():
                votes[t][d].append((v, s_name))

    final_test_data = defaultdict(dict)
    for t, dv in votes.items():
        for d, vlist in dv.items():
            if not vlist:
                continue
            if t in ["Dengue NS1 Antigen", "Dengue IgG", "Dengue IgM"]:
                dedicated = [v for v, s in vlist if s == "Dengue-Dedicated"]
                if dedicated:
                    final_test_data[t][d] = dedicated[0]
                    continue
            c = Counter([v for v, _ in vlist])
            val, count = c.most_common(1)[0]
            if count / len(vlist) >= 0.4:
                final_test_data[t][d] = val

    return pdf.name, content.content_hash if content else None, meta, dict(final_test_data)


def run_transcribe(
    input_dir: Path,
    output_dir: Path,
    job_name: str,
    verbose: bool = False,
    threads: int = os.cpu_count(),
):
    """
    Phase 1: PDF text extraction, test mapping, and QC hashing.
    Saves intermediate data to disk for downstream subcommands.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logging(output_dir, job_name, verbose=verbose)
    logger.info(f"Input: {input_dir}")
    logger.info(f"Output: {output_dir}")

    pdf_files = list(input_dir.glob("*.pdf"))
    if not pdf_files:
        logger.error("No PDFs found.")
        print("No PDFs found in input directory.")
        return

    all_extractions = []
    failed_files = []
    seen_hashes = {}
    qc_duplicates = []  # List of dicts: [{'duplicate': 'A.pdf', 'original': 'B.pdf'}, ...]

    logger.info(f"Found {len(pdf_files)} PDFs. Starting extraction...")
    print(f"Starting extraction for {len(pdf_files)} files...")

    start_time = time.time()

    if not verbose:
        with alive_bar(len(pdf_files), title="Transcribing PDFs", force_tty=True) as bar:
            with concurrent.futures.ProcessPoolExecutor(max_workers=threads) as executor:
                futures = {
                    executor.submit(_process_pdf_worker, pdf, output_dir, job_name, verbose): pdf
                    for pdf in pdf_files
                }
                for future in concurrent.futures.as_completed(futures):
                    pdf_name, content_hash, meta, final_test_data = future.result()
                    if not content_hash:
                        failed_files.append(pdf_name)
                    else:
                        if content_hash in seen_hashes:
                            orig = seen_hashes[content_hash]
                            qc_duplicates.append({"duplicate": pdf_name, "original": orig})
                        else:
                            seen_hashes[content_hash] = pdf_name
                        if final_test_data:
                            all_extractions.append((pdf_name, meta, final_test_data))
                        else:
                            failed_files.append(pdf_name)
                    bar()
    else:
        # Sequential processing for verbose mode to keep logs readable and in order
        for idx, pdf in enumerate(pdf_files, 1):
            pdf_name, content_hash, meta, final_test_data = _process_pdf_worker(
                pdf, output_dir, job_name, verbose
            )
            if not content_hash:
                failed_files.append(pdf_name)
            else:
                if content_hash in seen_hashes:
                    orig = seen_hashes[content_hash]
                    qc_duplicates.append({"duplicate": pdf_name, "original": orig})
                else:
                    seen_hashes[content_hash] = pdf_name
                if final_test_data:
                    all_extractions.append((pdf_name, meta, final_test_data))
                else:
                    failed_files.append(pdf_name)

    # Aggregate
    aggregated = _aggregate_data(all_extractions)

    # Identify longitudinal
    patient_data = aggregated["patient_data"]
    longitudinal_tests = identify_longitudinal_tests(patient_data)

    # Save intermediate data
    payload = {
        "all_extractions": all_extractions,
        "failed_files": failed_files,
        "qc_duplicates": qc_duplicates,
        "longitudinal_tests": longitudinal_tests,
        **aggregated,
    }
    save_extraction_data(output_dir, payload, job_name)

    elapsed_time = time.time() - start_time
    minutes, seconds = divmod(int(elapsed_time), 60)

    manual_time_minutes = len(pdf_files) * 5

    print(f"\nTotal time taken to transcribe: {minutes} Minutes and {seconds} Seconds.")
    print(
        f"Doing this manually would have taken you approximately {manual_time_minutes} minutes. You're welcome! 😉"
    )

    print(
        f"\nExtraction complete. {len(all_extractions)} files processed, {len(failed_files)} failed."
    )
    print("Check extraction.log for details.")
    logger.info("Extraction phase complete.")


# =============================================================================
# QC PHASE
# =============================================================================


def run_qc(output_dir: Path, job_name: str):
    """
    Phase 2: Report content hashing and duplicate detection results.
    """
    data = load_extraction_data(output_dir, job_name)
    qc_duplicates = data.get("qc_duplicates", [])
    failed_files = data.get("failed_files", [])

    print("\n" + "=" * 50)
    print("QUALITY CONTROL REPORT")
    print("=" * 50)
    print(f"Duplicate Content Files : {len(qc_duplicates)}")
    if qc_duplicates:
        for item in qc_duplicates:
            print(f"    ↳ {item['duplicate']} (identical to {item['original']})")
    print(f"Failed/No Data Files    : {len(failed_files)}")
    if failed_files:
        for f in failed_files:
            print(f"    ↳ {f}")
    print("=" * 50)


# =============================================================================
# STATS PHASE
# =============================================================================


def run_stats(output_dir: Path, job_name: str):
    """
    Phase 3: Generate the Stats_Refined.txt statistical report.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logging(output_dir, job_name, append=True)

    data = load_extraction_data(output_dir, job_name)

    generate_condensed_stats(
        latest_data=data["latest_data"],
        metadata_map=data["m_map"],
        all_extractions=data["all_extractions"],
        qc_dupes=data["qc_duplicates"],
        failed_files=data["failed_files"],
        file_name_map=data["f_map"],
        longitudinal_tests=data.get("longitudinal_tests", {}),
        patient_data=data.get("patient_data", {}),
        output_path=output_dir / f"{job_name}_Stats_Refined.txt",
        logger=logger,
    )
    print(f"Stats report saved to {output_dir / f'{job_name}_Stats_Refined.txt'}")


# =============================================================================
# PLOT PHASE
# =============================================================================


def run_plot(output_dir: Path, job_name: str):
    """
    Phase 4: Generate clinical distribution graphs.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logging(output_dir, job_name, append=True)

    data = load_extraction_data(output_dir, job_name)

    # Stratification Prompt
    do_strat = input("\nDo you want Mild, Moderate, Severe Stratification? (Y/N): ").strip().lower()
    if do_strat == "y":
        stratified_data, groups = perform_stratification(data["latest_data"], data["m_map"])

        # Stats
        stats_res = run_kruskal_wallis(stratified_data)
        print("\n--- Kruskal-Wallis Test Results ---")
        for var, res in stats_res.items():
            print(f"{var}: H={res['H-statistic']:.2f}, p-value={res['p-value']:.4f}")

        # Plot preference
        plot_type = input(
            "\nWhich plot do you prefer for visualization? (1) Box-and-Whisker (2) Violin Plot: "
        ).strip()
        pzfx_path = output_dir / f"{job_name}.pzfx"
        generate_pzfx(stratified_data, pzfx_path, job_name, plot_type)
        print(f"\nStratification complete. GraphPad Prism file saved to {pzfx_path}\n")

    generate_clinical_plots(
        latest_data=data["latest_data"],
        metadata_map=data["m_map"],
        output_dir=output_dir / f"{job_name}_Graphs",
        logger=logger,
    )
    print(f"Clinical plots saved to {output_dir / f'{job_name}_Graphs'}")


# =============================================================================
# FULL PIPELINE
# =============================================================================


def run_all_phase1(
    input_dir: Path,
    output_dir: Path,
    job_name: str,
    verbose: bool = False,
    threads: int = os.cpu_count(),
):
    """
    Phase 1 of the full pipeline: Transcribe + generate Master_Data.xlsx.
    Called before the verification gate.
    """
    # Transcribe
    run_transcribe(input_dir, output_dir, job_name, verbose=verbose, threads=threads)

    # Load aggregated data
    data = load_extraction_data(output_dir, job_name)
    logger = setup_logging(output_dir, job_name, append=True, verbose=verbose)

    # Generate Master Excel
    _generate_excel_report(data, output_dir, logger, job_name)

    # Generate Longitudinal Excel if data exists
    longitudinal_tests = data.get("longitudinal_tests", {})
    if longitudinal_tests:
        generate_longitudinal_excel(
            data["patient_data"],
            longitudinal_tests,
            output_dir / f"{job_name}_longitudinal_data.xlsx",
            logger,
        )

    logger.info("Phase 1 complete: Transcription and Excel reports generated.")


def run_all_phase2(output_dir: Path, job_name: str):
    """
    Phase 2 of the full pipeline: QC + Stats + Plot.
    Called after the verification gate approves continuation.
    """
    data = load_extraction_data(output_dir, job_name)
    logger = setup_logging(output_dir, job_name, append=True)

    # QC (terminal output)
    run_qc(output_dir, job_name)

    # Stats (pass longitudinal info)
    generate_condensed_stats(
        latest_data=data["latest_data"],
        metadata_map=data["m_map"],
        all_extractions=data["all_extractions"],
        qc_dupes=data["qc_duplicates"],
        failed_files=data["failed_files"],
        file_name_map=data["f_map"],
        longitudinal_tests=data.get("longitudinal_tests", {}),
        patient_data=data.get("patient_data", {}),
        output_path=output_dir / f"{job_name}_Stats_Refined.txt",
        logger=logger,
    )

    # Plot

    do_strat = input("\nDo you want Mild, Moderate, Severe Stratification? (Y/N): ").strip().lower()
    if do_strat == "y":
        stratified_data, groups = perform_stratification(data["latest_data"], data["m_map"])

        # Stats
        stats_res = run_kruskal_wallis(stratified_data)
        logger.info("--- Kruskal-Wallis Test Results ---")
        for var, res in stats_res.items():
            logger.info(f"{var}: H={res['H-statistic']:.2f}, p-value={res['p-value']:.4f}")

        # Plot preference
        plot_type = input(
            "\nWhich plot do you prefer for visualization? (1) Box-and-Whisker (2) Violin Plot: "
        ).strip()

        # ── Excel export (Mild / Moderate / Clinically Severe) ──────────
        from .r_integration import export_stratified_excel, generate_r_script, run_r_script

        excel_path = output_dir / f"{job_name}_mild_mod_severe.xlsx"
        export_stratified_excel(stratified_data, excel_path)
        logger.info(f"Stratified Excel saved to {excel_path}")

        # ── R script: Kruskal-Wallis + ggplot2 plots ────────────────────
        r_script_path = output_dir / f"{job_name}_analysis.R"
        generate_r_script(stratified_data, excel_path, r_script_path, plot_type)
        logger.info(f"R analysis script written to {r_script_path}")

        # Attempt to execute the R script automatically
        if run_r_script(r_script_path, logger):
            logger.info("R plots generated successfully (PNG + SVG).")
        else:
            logger.info(f'Run the script manually:  Rscript "{r_script_path}"')

        # ── GraphPad Prism (secondary, best-effort) ─────────────────────
        try:
            pzfx_path = output_dir / f"{job_name}.pzfx"
            generate_pzfx(stratified_data, pzfx_path, job_name, plot_type)
            logger.info(f"GraphPad Prism file saved to {pzfx_path}")
        except Exception as e:
            logger.warning(f"GraphPad PZFX generation skipped: {e}")

    generate_clinical_plots(
        latest_data=data["latest_data"],
        metadata_map=data["m_map"],
        output_dir=output_dir / f"{job_name}_Graphs",
        logger=logger,
    )

    logger.info("DONE.")
    print(f"\nAll reports saved to {output_dir}")


# =============================================================================
# HELPERS
# =============================================================================


def _aggregate_data(all_extractions: List) -> dict:
    """Aggregate per-PDF extractions into per-patient data."""
    p_data = defaultdict(lambda: defaultdict(dict))
    f_map = defaultdict(set)
    m_map = {}

    for fname, meta, tdata in all_extractions:
        mid = meta.get("MAX_id") or f"UNKNOWN_{fname}"
        f_map[mid].add(fname)
        if mid not in m_map or (not m_map[mid].get("Age") and meta.get("Age")):
            m_map[mid] = meta

        for t, dv in tdata.items():
            for d, v in dv.items():
                p_data[mid][t][d] = v

    latest_data = {}
    for mid, tests in p_data.items():
        latest_data[mid] = {}
        last_dt = None
        for t, dv in tests.items():
            valid = []
            for dstr, val in dv.items():
                try:
                    dt = datetime.strptime(dstr, "%d-%m-%Y")
                    valid.append((dt, dstr, val))
                except Exception:
                    pass
            if valid:
                valid.sort(reverse=True)
                latest_data[mid][t] = valid[0][2]
                if not last_dt or valid[0][0] > last_dt:
                    last_dt = valid[0][0]
                    latest_data[mid]["collection_date"] = valid[0][1]
            else:
                latest_data[mid][t] = "nil"

    return {
        "patient_data": dict(p_data),
        "latest_data": latest_data,
        "f_map": dict(f_map),
        "m_map": m_map,
    }


def identify_longitudinal_tests(patient_data: Dict) -> Dict:
    """
    Identify tests with MULTIPLE dates (>= 2 distinct dates).
    Refined Definition: Any MaxID that has at least one identical clinical test
    recorded on two or more different Collection Dates.
    """
    longitudinal = {}

    for max_id, tests in patient_data.items():
        multi_date = []
        for test_name, date_values in tests.items():
            # Get valid distinct dates
            raw_dates = set([d for d in date_values.keys() if d != "nil"])

            # Simple check: Are there 2 or more distinct dates?
            if len(raw_dates) >= 2:
                multi_date.append(test_name)

        if multi_date:
            longitudinal[max_id] = multi_date

    return longitudinal


def _generate_excel_report(data: dict, output_dir: Path, logger: logging.Logger, job_name: str):
    """Generate Master_Data_Refined.xlsx with branding and highlighting."""
    latest_data = data["latest_data"]
    f_map = data["f_map"]
    m_map = data["m_map"]
    longitudinal_tests = data.get("longitudinal_tests", {})

    rows = []
    # Sort by MAX_id for consistency
    sorted_mids = sorted(latest_data.keys())

    for mid in sorted_mids:
        row = {
            "MAX_id": mid,
            "Tests Done": determine_tests_done(latest_data[mid]),
            "File Name": ", ".join(sorted(list(f_map[mid]))),
            "SIN No.": m_map.get(mid, {}).get("SIN_No", "nil"),
            "Gender": m_map.get(mid, {}).get("Gender", "nil"),
            "Age": m_map.get(mid, {}).get("Age", "nil"),
            "Collection_date": latest_data[mid].get("collection_date", "nil"),
        }
        for c in OUTPUT_COLUMNS[7:]:
            row[c] = latest_data[mid].get(c, "nil")
        rows.append(row)

    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    xlsx_path = output_dir / f"{job_name}_Master_Data_Refined.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Master")
        ws = writer.sheets["Master"]

        # Styles
        header_font = Font(bold=True)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Bold Headers
        for cell in ws[1]:
            cell.font = header_font

        # Highlight Longitudinal MAX_id (Column A)
        # rows are 1-indexed. Header is row 1. Data starts row 2.
        for row_idx, mid in enumerate(df["MAX_id"], start=2):
            if mid in longitudinal_tests:
                ws.cell(row=row_idx, column=1).fill = yellow_fill

        # Highlight Dengue Positives
        dengue_thresholds = {
            "Dengue NS1 Antigen": 0.9,
            "Dengue IgG": 9.0,
            "Dengue IgM": 9.0,
        }

        # Map column names to indices (1-based)
        col_map = {name: idx for idx, name in enumerate(OUTPUT_COLUMNS, start=1)}

        for col_name, threshold in dengue_thresholds.items():
            if col_name in col_map:
                col_idx = col_map[col_name]
                for row_idx in range(2, len(df) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val and val != "nil" and str(val).strip():
                        try:
                            # Clean string value to float
                            clean_val = str(val).lower().replace(">", "").replace("<", "").strip()
                            if clean_val in ["positive", "reactive", "detected"]:
                                cell.fill = yellow_fill
                            else:
                                fval = float(clean_val)
                                if fval >= threshold:
                                    cell.fill = yellow_fill
                        except ValueError:
                            pass

    logger.info(f"Excel report saved: {xlsx_path}")


def generate_longitudinal_excel(
    patient_data: Dict, longitudinal_tests: Dict, output_path: Path, logger: logging.Logger
):
    """Generate longitudinal_data.xlsx for patients with multi-date records."""
    rows = []

    for max_id, test_names in longitudinal_tests.items():
        for test_name in test_names:
            date_values = patient_data[max_id][test_name]

            sorted_items = []
            for date_str, value in date_values.items():
                try:
                    dt = datetime.strptime(date_str, "%d-%m-%Y")
                    sorted_items.append((dt, date_str, value))
                except ValueError:
                    pass

            sorted_items.sort()  # sort by date ascending

            for _, date_str, value in sorted_items:
                rows.append(
                    {
                        "MAX_id": max_id,
                        "Test_Name": test_name,
                        "Collection_Date": date_str,
                        "Value": value,
                    }
                )

    if not rows:
        return

    df = pd.DataFrame(rows, columns=["MAX_id", "Test_Name", "Collection_Date", "Value"])
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Longitudinal")
        ws = writer.sheets["Longitudinal"]
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
    logger.info(f"Longitudinal report saved: {output_path} ({len(rows)} records)")
