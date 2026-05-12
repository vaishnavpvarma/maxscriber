"""
MaXScriber v1.0
Pipeline orchestration: logging, extraction, aggregation, and reporting.
Supports both full pipeline and individual subcommand execution.
"""

import sys
import logging
import pickle
from pathlib import Path
from typing import Dict, List, Tuple
from collections import defaultdict, Counter
from datetime import datetime, timedelta
import time

import pandas as pd
from openpyxl.styles import Font, PatternFill

from maxscriber.constants import OUTPUT_COLUMNS
from maxscriber.extraction import (
    extract_pdf_content,
    extract_metadata,
    extract_dates_from_content,
    func_table_smart,
    func_line_pattern,
    func_dengue_dedicated,
    func_kft_table,
    determine_tests_done,
)
from maxscriber.stats import generate_condensed_stats
from maxscriber.plots import generate_clinical_plots


# =============================================================================
# INTERMEDIATE DATA PERSISTENCE
# =============================================================================

DATA_FILE = 'extraction_data.pkl'


def save_extraction_data(output_dir: Path, payload: dict):
    """Save intermediate extraction data for subcommand independence."""
    with open(output_dir / DATA_FILE, 'wb') as f:
        pickle.dump(payload, f)


def load_extraction_data(output_dir: Path) -> dict:
    """Load previously saved extraction data."""
    pkl_path = output_dir / DATA_FILE
    if not pkl_path.exists():
        print(f"ERROR: No extraction data found at {pkl_path}")
        print("Run 'maxscriber transcribe' first to generate extraction data.")
        sys.exit(1)
    with open(pkl_path, 'rb') as f:
        return pickle.load(f)


# =============================================================================
# LOGGING
# =============================================================================

def setup_logging(output_dir: Path, append: bool = False) -> logging.Logger:
    """Setup comprehensive logging (V2 style).
    
    Args:
        output_dir: Directory for the log file.
        append: If True, append to existing log (for Phase 2). 
                If False, start fresh (for Phase 1).
    """
    log_file = output_dir / 'extraction.log'

    # Clear any existing handlers to avoid duplicate output on re-runs
    root = logging.getLogger()
    for h in root.handlers[:]:
        root.removeHandler(h)

    file_mode = 'a' if append else 'w'
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[
            logging.FileHandler(log_file, mode=file_mode),
            logging.StreamHandler(sys.stdout),
        ],
    )
    logger = logging.getLogger('MaxScriber')

    if append:
        logger.info("")
        logger.info("=" * 80)
        logger.info("Resuming pipeline — Phase 2 (QC / Stats / Plot)")
        logger.info("=" * 80)
    else:
        # Banner in log
        logger.info("=" * 80)
        logger.info("Intelligent Multi-Pass Medical PDF Extractor")
        logger.info("No AI/ML - Pure Rule-Based with Cross-Validation")
        logger.info("=" * 80)
    return logger


# =============================================================================
# EXTRACTION PHASE
# =============================================================================

def run_transcribe(input_dir: Path, output_dir: Path):
    """
    Phase 1: PDF text extraction, test mapping, and QC hashing.
    Saves intermediate data to disk for downstream subcommands.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logging(output_dir)
    logger.info(f"Input: {input_dir}")
    logger.info(f"Output: {output_dir}")

    pdf_files = list(input_dir.glob('*.pdf'))
    if not pdf_files:
        logger.error("No PDFs found.")
        print("No PDFs found in input directory.")
        return

    all_extractions = []
    failed_files = []
    seen_hashes = {}
    qc_duplicates = []  # List of dicts: [{'duplicate': 'A.pdf', 'original': 'B.pdf'}, ...]

    logger.info(f"Found {len(pdf_files)} PDFs. Starting extraction...")
    print(f"Starting extraction for {len(pdf_files)} files...")
    
    start_time = time.time()

    for idx, pdf in enumerate(pdf_files, 1):
        logger.info(f"\n{'=' * 80}")
        logger.info(f"PDF {idx}/{len(pdf_files)}: {pdf.name}")
        logger.info(f"{'=' * 80}")

        content = extract_pdf_content(str(pdf), logger)

        if not content:
            failed_files.append(pdf.name)
            continue

        # QC: Check Duplicate Content
        if content.content_hash in seen_hashes:
            orig = seen_hashes[content.content_hash]
            logger.warning(f"DUPLICATE DETECTED: {pdf.name} is identical to {orig}")
            qc_duplicates.append({'duplicate': pdf.name, 'original': orig})
        else:
            seen_hashes[content.content_hash] = pdf.name

        meta = extract_metadata(content, logger)
        dates = extract_dates_from_content(content, logger)

        res_table = func_table_smart(content, dates, logger)
        res_line = func_line_pattern(content, dates, logger)
        res_dengue = func_dengue_dedicated(content, dates, logger)
        res_kft = func_kft_table(content, dates, logger)

        # Pull out the KFT unit sidecar before merging into votes
        kft_units = res_kft.pop('__kft_units__', {})

        # Strategies & Voting
        votes = defaultdict(lambda: defaultdict(list))
        strategies = [
            ("Table", res_table),
            ("Line", res_line),
            ("Dengue-Dedicated", res_dengue),
            ("KFT-Table", res_kft),
        ]

        for s_name, res in strategies:
            for t, dv in res.items():
                for d, v in dv.items():
                    votes[t][d].append((v, s_name))

        final_test_data = defaultdict(dict)
        for t, dv in votes.items():
            for d, vlist in dv.items():
                if not vlist:
                    continue
                # Dengue Priority
                if t in ['Dengue NS1 Antigen', 'Dengue IgG', 'Dengue IgM']:
                    dedicated = [v for v, s in vlist if s == 'Dengue-Dedicated']
                    if dedicated:
                        final_test_data[t][d] = dedicated[0]
                        continue
                # Voting
                c = Counter([v for v, _ in vlist])
                val, count = c.most_common(1)[0]
                if count / len(vlist) >= 0.4:
                    final_test_data[t][d] = val

        if final_test_data:
            all_extractions.append((pdf.name, meta, dict(final_test_data), kft_units))
        else:
            failed_files.append(pdf.name)
            logger.warning(f"No data extracted from {pdf.name}")

    # Aggregate
    aggregated = _aggregate_data(all_extractions)
    
    # Identify longitudinal
    patient_data = aggregated['patient_data']
    longitudinal_tests = identify_longitudinal_tests(patient_data)

    # Save intermediate data
    payload = {
        'all_extractions': all_extractions,
        'failed_files': failed_files,
        'qc_duplicates': qc_duplicates,
        'longitudinal_tests': longitudinal_tests,
        **aggregated,
    }
    save_extraction_data(output_dir, payload)

    elapsed_time = time.time() - start_time
    minutes, seconds = divmod(int(elapsed_time), 60)
    
    manual_time_minutes = len(pdf_files) * 5
    
    print(f"\nTotal time taken to transcribe: {minutes} Minutes and {seconds} Seconds.")
    print(f"Doing this manually would have taken you approximately {manual_time_minutes} minutes. You're welcome! ;)")
    
    print(f"\nExtraction complete. {len(all_extractions)} files processed, {len(failed_files)} failed.")
    print(f"Check extraction.log for details.")
    logger.info("Extraction phase complete.")


# =============================================================================
# QC PHASE
# =============================================================================

def run_qc(output_dir: Path):
    """
    Phase 2: Report content hashing and duplicate detection results.
    """
    data = load_extraction_data(output_dir)
    qc_duplicates = data.get('qc_duplicates', [])
    failed_files = data.get('failed_files', [])

    print("\n" + "=" * 50)
    print("QUALITY CONTROL REPORT")
    print("=" * 50)
    print(f"Duplicate Content Files : {len(qc_duplicates)}")
    if qc_duplicates:
        for item in qc_duplicates:
            print(f"    ↳ {item['duplicate']} (identical to {item['original']})")
    print(f"Failed/No Data Files    : {len(failed_files)}")
    if failed_files:
        for f in failed_files:
            print(f"    ↳ {f}")
    print("=" * 50)


# =============================================================================
# STATS PHASE
# =============================================================================

def run_stats(output_dir: Path):
    """
    Phase 3: Generate the Stats_Refined.txt statistical report.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logging(output_dir, append=True)

    data = load_extraction_data(output_dir)

    generate_condensed_stats(
        latest_data=data['latest_data'],
        metadata_map=data['m_map'],
        all_extractions=data['all_extractions'],
        qc_dupes=data['qc_duplicates'],
        failed_files=data['failed_files'],
        file_name_map=data['f_map'],
        longitudinal_tests=data.get('longitudinal_tests', {}),
        patient_data=data.get('patient_data', {}),
        output_path=output_dir / 'Stats_Refined.txt',
        logger=logger,
    )
    print(f"Stats report saved to {output_dir / 'Stats_Refined.txt'}")


# =============================================================================
# PLOT PHASE
# =============================================================================

def run_plot(output_dir: Path):
    """
    Phase 4: Generate clinical distribution graphs.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logging(output_dir, append=True)

    data = load_extraction_data(output_dir)

    generate_clinical_plots(
        latest_data=data['latest_data'],
        metadata_map=data['m_map'],
        output_dir=output_dir,
        logger=logger,
    )
    print(f"Clinical plots saved to {output_dir / 'graphs'}")


# =============================================================================
# FULL PIPELINE
# =============================================================================

def run_all_phase1(input_dir: Path, output_dir: Path):
    """
    Phase 1 of the full pipeline: Transcribe + generate Master_Data.xlsx.
    Called before the verification gate.
    """
    # Transcribe
    run_transcribe(input_dir, output_dir)

    # Load aggregated data and generate Excel report
    # Load aggregated data
    data = load_extraction_data(output_dir)
    logger = setup_logging(output_dir, append=True)
    
    # Generate Master Excel
    _generate_excel_report(data, output_dir, logger)
    
    # Generate Longitudinal Excel if data exists
    longitudinal_tests = data.get('longitudinal_tests', {})
    if longitudinal_tests:
        generate_longitudinal_excel(
            data['patient_data'], 
            longitudinal_tests, 
            output_dir / 'longitudinal_data.xlsx', 
            logger
        )

    # Generate Lab Variance Report
    generate_lab_variance_report(
        centre_map=data.get('centre_map', {}),
        patient_data=data.get('patient_data', {}),
        f_map=data.get('f_map', {}),
        kft_unit_map=data.get('kft_unit_map', {}),
        output_path=output_dir / 'lab_variance_report.txt',
        logger=logger,
    )
    
    logger.info("Phase 1 complete: Transcription and Excel reports generated.")


def run_all_phase2(output_dir: Path):
    """
    Phase 2 of the full pipeline: QC + Stats + Plot.
    Called after the verification gate approves continuation.
    """
    data = load_extraction_data(output_dir)
    logger = setup_logging(output_dir, append=True)

    # QC (terminal output)
    run_qc(output_dir)

    # Stats (pass longitudinal info)
    generate_condensed_stats(
        latest_data=data['latest_data'],
        metadata_map=data['m_map'],
        all_extractions=data['all_extractions'],
        qc_dupes=data['qc_duplicates'],
        failed_files=data['failed_files'],
        file_name_map=data['f_map'],
        longitudinal_tests=data.get('longitudinal_tests', {}),
        patient_data=data.get('patient_data', {}),
        output_path=output_dir / 'Stats_Refined.txt',
        logger=logger,
    )

    # Plot
    generate_clinical_plots(
        latest_data=data['latest_data'],
        metadata_map=data['m_map'],
        output_dir=output_dir,
        logger=logger,
    )

    logger.info("DONE.")
    print(f"\nAll reports saved to {output_dir}")


# =============================================================================
# HELPERS
# =============================================================================

def _aggregate_data(all_extractions: List) -> dict:
    """Aggregate per-PDF extractions into per-patient data."""
    p_data = defaultdict(lambda: defaultdict(dict))
    f_map = defaultdict(set)
    m_map = {}
    # centre_map: MAX_id -> {date -> {centre, fname}} for variance detection
    centre_map: Dict[str, List[dict]] = defaultdict(list)
    # kft_unit_map: MAX_id -> test -> {date -> unit}
    kft_unit_map: Dict[str, Dict[str, Dict[str, str]]] = defaultdict(lambda: defaultdict(dict))

    for record in all_extractions:
        # Support both old 3-tuple (backward compat) and new 4-tuple
        if len(record) == 4:
            fname, meta, tdata, kft_units = record
        else:
            fname, meta, tdata = record
            kft_units = {}

        mid = meta.get('MAX_id') or f"UNKNOWN_{fname}"
        f_map[mid].add(fname)
        if mid not in m_map or (not m_map[mid].get('Age') and meta.get('Age')):
            m_map[mid] = meta

        # Track centre per date for variance reporting
        centre = meta.get('Centre')
        for t, dv in tdata.items():
            for d, v in dv.items():
                p_data[mid][t][d] = v
                if centre:
                    centre_map[mid].append({
                        'date': d,
                        'centre': centre,
                        'fname': fname,
                    })

        # Merge KFT units
        for test_nm, date_unit in kft_units.items():
            for d, u in date_unit.items():
                kft_unit_map[mid][test_nm][d] = u

    latest_data = {}
    for mid, tests in p_data.items():
        latest_data[mid] = {}
        last_dt = None
        for t, dv in tests.items():
            valid = []
            for dstr, val in dv.items():
                try:
                    dt = datetime.strptime(dstr, '%d-%m-%Y')
                    valid.append((dt, dstr, val))
                except Exception:
                    pass
            if valid:
                valid.sort(reverse=True)
                latest_data[mid][t] = valid[0][2]
                if not last_dt or valid[0][0] > last_dt:
                    last_dt = valid[0][0]
                    latest_data[mid]['collection_date'] = valid[0][1]
            else:
                latest_data[mid][t] = 'nil'

    return {
        'patient_data': dict(p_data),
        'latest_data': latest_data,
        'f_map': dict(f_map),
        'm_map': m_map,
        'centre_map': dict(centre_map),
        'kft_unit_map': dict(kft_unit_map),
    }


def identify_longitudinal_tests(patient_data: Dict) -> Dict:
    """
    Identify tests with MULTIPLE dates (>= 2 distinct dates).
    Refined Definition: Any MaxID that has at least one identical clinical test 
    recorded on two or more different Collection Dates.
    """
    longitudinal = {}

    for max_id, tests in patient_data.items():
        multi_date = []
        for test_name, date_values in tests.items():
            # Get valid distinct dates
            raw_dates = set([d for d in date_values.keys() if d != 'nil'])
            
            # Simple check: Are there 2 or more distinct dates?
            if len(raw_dates) >= 2:
                multi_date.append(test_name)

        if multi_date:
            longitudinal[max_id] = multi_date

    return longitudinal


def _generate_excel_report(data: dict, output_dir: Path, logger: logging.Logger):
    """Generate Master_Data_Refined.xlsx with branding and highlighting."""
    latest_data = data['latest_data']
    f_map = data['f_map']
    m_map = data['m_map']
    longitudinal_tests = data.get('longitudinal_tests', {})

    rows = []
    # Sort by MAX_id for consistency
    sorted_mids = sorted(latest_data.keys())
    
    for mid in sorted_mids:
        row = {
            'MAX_id': mid,
            'Centre': m_map.get(mid, {}).get('Centre', 'nil'),
            'Tests Done': determine_tests_done(latest_data[mid]),
            'File Name': ', '.join(sorted(list(f_map[mid]))),
            'SIN No.': m_map.get(mid, {}).get('SIN_No', 'nil'),
            'Gender': m_map.get(mid, {}).get('Gender', 'nil'),
            'Age': m_map.get(mid, {}).get('Age', 'nil'),
            'Collection_date': latest_data[mid].get('collection_date', 'nil'),
        }
        from maxscriber.constants import TEST_ALIASES
        for c in list(TEST_ALIASES.keys()):
            row[c] = latest_data[mid].get(c, 'nil')
            # Extract corresponding unit if tracking
            col_date = latest_data[mid].get('collection_date')
            unit_map = data.get('kft_unit_map', {})
            unit_val = unit_map.get(mid, {}).get(c, {}).get(col_date, 'nil') if col_date else 'nil'
            row[f"{c}_Unit"] = unit_val
        rows.append(row)

    df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    xlsx_path = output_dir / 'Master_Data_Refined.xlsx'
    
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Master')
        ws = writer.sheets['Master']
        
        # Styles
        header_font = Font(bold=True)
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Bold Headers
        for cell in ws[1]:
            cell.font = header_font

        # Highlight Longitudinal MAX_id (Column A)
        # rows are 1-indexed. Header is row 1. Data starts row 2.
        for row_idx, mid in enumerate(df['MAX_id'], start=2):
            if mid in longitudinal_tests:
                ws.cell(row=row_idx, column=1).fill = yellow_fill

        # Highlight Dengue Positives
        dengue_thresholds = {
            'Dengue NS1 Antigen': 0.9,
            'Dengue IgG': 9.0,
            'Dengue IgM': 9.0,
        }
        
        # Map column names to indices (1-based)
        col_map = {name: idx for idx, name in enumerate(OUTPUT_COLUMNS, start=1)}
        
        for col_name, threshold in dengue_thresholds.items():
            if col_name in col_map:
                col_idx = col_map[col_name]
                for row_idx in range(2, len(df) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val and val != 'nil' and str(val).strip():
                        try:
                            # Clean string value to float
                            clean_val = str(val).lower().replace('>', '').replace('<', '').strip()
                            if clean_val in ['positive', 'reactive', 'detected']:
                                cell.fill = yellow_fill
                            else:
                                fval = float(clean_val)
                                if fval >= threshold:
                                    cell.fill = yellow_fill
                        except ValueError:
                            pass

    logger.info(f"Excel report saved: {xlsx_path}")


def generate_longitudinal_excel(patient_data: Dict, longitudinal_tests: Dict, output_path: Path, logger: logging.Logger):
    """Generate longitudinal_data.xlsx for patients with multi-date records."""
    rows = []

    for max_id, test_names in longitudinal_tests.items():
        for test_name in test_names:
            date_values = patient_data[max_id][test_name]

            sorted_items = []
            for date_str, value in date_values.items():
                try:
                    dt = datetime.strptime(date_str, '%d-%m-%Y')
                    sorted_items.append((dt, date_str, value))
                except ValueError:
                    pass

            sorted_items.sort() # sort by date ascending

            for _, date_str, value in sorted_items:
                rows.append({
                    'MAX_id': max_id,
                    'Test_Name': test_name,
                    'Collection_Date': date_str,
                    'Value': value
                })

    if not rows:
        return

    df = pd.DataFrame(rows, columns=['MAX_id', 'Test_Name', 'Collection_Date', 'Value'])
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Longitudinal')
        ws = writer.sheets['Longitudinal']
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
    logger.info(f"Longitudinal report saved: {output_path} ({len(rows)} records)")

# =============================================================================
# LAB VARIANCE REPORT
# =============================================================================

def generate_lab_variance_report(
    centre_map: Dict,
    patient_data: Dict,
    f_map: Dict,
    kft_unit_map: Dict,
    output_path: Path,
    logger: logging.Logger,
):
    """
    Detect patients whose KFT results were obtained from multiple lab centres,
    then emit a structured lab_variance_report.txt.

    A Lab-Based Change is flagged when the same MAX_ID has entries linked to
    two or more different Centre values, indicating:
      - A shift in testing centre between visits, OR
      - Differing lab methodology or units for the same test.
    """
    from maxscriber.constants import KFT_TESTS
    from datetime import datetime as _dt

    flagged = []

    for max_id, entries in centre_map.items():
        # Deduplicate by (date, centre, fname)
        seen = set()
        unique_entries = []
        for e in entries:
            key = (e["date"], e["centre"], e["fname"])
            if key not in seen:
                seen.add(key)
                unique_entries.append(e)

        # Distinct centres for this patient
        centres = {e["centre"] for e in unique_entries}
        if len(centres) < 2:
            continue  # Same centre for all visits

        detail_lines = []
        tests_examined = patient_data.get(max_id, {})

        for test_name in sorted(KFT_TESTS):
            if test_name not in tests_examined:
                continue
            date_values = tests_examined[test_name]
            for date_str, value in sorted(date_values.items()):
                matching = [e for e in unique_entries if e["date"] == date_str]
                centre_label = matching[0]["centre"] if matching else "Unknown"
                fname_label = matching[0]["fname"] if matching else "Unknown"
                unit = (kft_unit_map
                        .get(max_id, {})
                        .get(test_name, {})
                        .get(date_str, ""))
                result_str = (f"{value} {unit}").strip()
                detail_lines.append(
                    f"  Date: {date_str} | Centre: {centre_label} | "
                    f"Test: {test_name} | Result: {result_str} | File: {fname_label}"
                )

        # Fallback if no KFT tests found
        if not detail_lines:
            for e in unique_entries:
                detail_lines.append(
                    "  Date: {} | Centre: {} | Test: (Non-KFT) | Result: N/A | File: {}".format(
                        e["date"], e["centre"], e["fname"]
                    )
                )

        # Determine observation type
        obs = "Shift in Testing Centre"
        for test_name in KFT_TESTS:
            units_seen = set(
                kft_unit_map.get(max_id, {}).get(test_name, {}).values()
            )
            if len(units_seen) > 1:
                obs = "Change in Methodology / Shift in Testing Centre"
                break

        source_pdfs = sorted(f_map.get(max_id, set()))
        flagged.append({
            "max_id": max_id,
            "observation": obs,
            "detail_lines": detail_lines,
            "source_pdfs": source_pdfs,
        })

    # Write report file
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        f.write("LAB VARIANCE & OBSERVATION REPORT\n")
        f.write("Generated: {}\n".format(_dt.now().strftime("%Y-%m-%d %H:%M:%S")))
        f.write("Total Flagged Patients: {}\n".format(len(flagged)))
        f.write("=" * 70 + "\n\n")

        if not flagged:
            f.write("No lab-based variances detected.\n")
        else:
            for entry in flagged:
                f.write("Patient ID  : {}\n".format(entry["max_id"]))
                f.write("Observation : {}\n".format(entry["observation"]))
                f.write("Details:\n")
                for line in entry["detail_lines"]:
                    f.write(line + "\n")
                f.write(
                    "Tracing     : Source PDFs: "
                    + ", ".join(entry["source_pdfs"]) + "\n"
                )
                f.write("-" * 70 + "\n\n")

    logger.info(
        "Lab variance report saved: {} ({} flagged patients)".format(
            output_path, len(flagged)
        )
    )
    print("Lab variance report saved to {} ({} flagged)".format(
        output_path, len(flagged)
    ))
