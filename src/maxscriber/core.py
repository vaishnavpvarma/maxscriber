"""
MaXScriber v1.0
Pipeline orchestration: logging, extraction, aggregation, and reporting.
Supports both full pipeline and individual subcommand execution.
"""

import sys
import os
import json
import sqlite3
import logging
import pickle
from pathlib import Path
from typing import Dict, List, Tuple
from collections import defaultdict, Counter
from datetime import datetime, timedelta
import time
import concurrent.futures
from alive_progress import alive_bar

import pandas as pd
from openpyxl.styles import Font, PatternFill

from maxscriber.constants import OUTPUT_COLUMNS
from maxscriber.extraction import (
    extract_pdf_content,
    extract_dates_from_content,
    determine_tests_done,
)
from maxscriber.adaptive_extractor import AdaptiveExtractor
from maxscriber.schema_manager import SchemaManager
from maxscriber.stats import generate_condensed_stats
from maxscriber.plots import generate_clinical_plots


# =============================================================================
# INTERMEDIATE DATA PERSISTENCE
# =============================================================================

def save_extraction_data(output_dir: Path, payload: dict, job_name: str = 'default'):
    """Save intermediate extraction data for subcommand independence."""
    pkl_name = f"{job_name}_extraction_data.pkl"
    with open(output_dir / pkl_name, 'wb') as f:
        pickle.dump(payload, f)


def load_extraction_data(output_dir: Path, job_name: str = 'default') -> dict:
    """Load previously saved extraction data."""
    pkl_name = f"{job_name}_extraction_data.pkl"
    pkl_path = output_dir / pkl_name
    if not pkl_path.exists():
        print(f"ERROR: No extraction data found at {pkl_path}")
        print(f"Run 'maxscriber transcribe --job-name {job_name}' first to generate extraction data.")
        sys.exit(1)
    with open(pkl_path, 'rb') as f:
        return pickle.load(f)


# =============================================================================
# LOGGING
# =============================================================================

def setup_logging(output_dir: Path, append: bool = False, verbose: bool = True, log_filename: str = 'extraction.log', job_name: str = 'default') -> logging.Logger:
    """Setup comprehensive logging (V2 style).
    
    Args:
        output_dir: Directory for the log file.
        append: If True, append to existing log.
        verbose: If True, log to stdout.
        log_filename: Filename for the log file.
        job_name: Prefix for the log file name.
    """
    if log_filename == 'extraction.log':
        log_file = output_dir / f"{job_name}_{log_filename}"
    else:
        log_file = output_dir / log_filename

    # Clear any existing handlers to avoid duplicate output on re-runs
    root = logging.getLogger(job_name)
    for h in root.handlers[:]:
        root.removeHandler(h)

    file_mode = 'a' if append else 'w'
    handlers = [
        logging.FileHandler(log_file, mode=file_mode, encoding='utf-8')
    ]
    if verbose:
        handlers.append(logging.StreamHandler(sys.stdout))

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=handlers,
    )
    logger = logging.getLogger(job_name)

    if append:
        if log_filename == 'extraction.log':
            logger.info("")
            logger.info("=" * 80)
            logger.info(f"Resuming pipeline — Phase 2 (QC / Stats / Plot) for {job_name}")
            logger.info("=" * 80)
    else:
        # Banner in log
        logger.info("=" * 80)
        logger.info(f"Intelligent Multi-Pass Medical PDF Extractor [Job: {job_name}]")
        logger.info("No AI/ML - Pure Rule-Based with Cross-Validation")
        logger.info("=" * 80)
    return logger


def _process_pdf_worker(pdf: Path, output_dir: Path, job_name: str, verbose: bool, pattern: str = '2024_2025', selected_categories: list = None, schema_name: str = 'MaxHospitals_Dengue') -> Tuple:
    """
    Process a single PDF file using the AdaptiveExtractor.
    """
    logger = setup_logging(output_dir, append=True, verbose=verbose, log_filename='extraction.log', job_name=job_name)
    try:
        import pdfplumber
        
        # Load schema
        sm = SchemaManager()
        schema = sm.get_schema(schema_name)
        if not schema:
            logger.error(f"Schema {schema_name} not found.")
            return pdf.name, None, {}, {}, {}
            
        extractor = AdaptiveExtractor(schema)
        
        # Extract content natively for adaptive extractor
        text_content = ""
        tables = []
        with pdfplumber.open(pdf) as p:
            for page in p.pages:
                page_text = page.extract_text()
                if page_text:
                    text_content += page_text + "\n"
                page_tables = page.extract_tables()
                if page_tables:
                    tables.extend(page_tables)

        meta = extractor.extract_metadata(text_content)
        final_test_data = extractor.extract_tests(tables, text_content)
        
        # For compatibility with legacy pipeline expectations
        content = extract_pdf_content(str(pdf), logger)
        content_hash = content.content_hash if content else None
        
        # We wrap the results by date for the legacy data structure
        # (Assuming today's date for simplicity in this bridge)
        # Ideally, we should parse dates using the collection date from meta
        collection_date = meta.get("Collection Date", "nil")
        if collection_date == "nil":
            collection_date = datetime.now().strftime("%d-%m-%Y")
            
        test_data_by_date = {t: {collection_date: v} for t, v in final_test_data.items() if v != 'nil'}
        kft_units = {} # Unit parsing to be added




        # Confidence heuristic
        total_tests = len(final_test_data)
        populated = sum(1 for v in final_test_data.values() if v != 'nil')
        confidence = (populated / total_tests * 100) if total_tests > 0 else 0.0

        return pdf.name, content_hash, meta, dict(test_data_by_date), kft_units, confidence
    except Exception as e:
        logger.error(f"Error processing {pdf.name}: {e}", exc_info=True)
        return pdf.name, None, {}, {}, {}, 0.0


# =============================================================================
# EXTRACTION PHASE
# =============================================================================

def _save_to_sqlite(output_dir: Path, payload: dict, job_name: str = 'default'):
    """Save aggregated extraction payload to a SQLite database."""
    db_path = output_dir / f'{job_name}_extractions.db'
    
    # Remove old database if exists to ensure clean run
    if db_path.exists():
        try:
            db_path.unlink()
        except Exception:
            pass

    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()

    # 1. Table for individual PDF extractions
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS extractions (
            pdf_name TEXT PRIMARY KEY,
            metadata TEXT,
            test_data TEXT,
            kft_units TEXT
        )
    """)

    # 2. Table for patients demographics
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS patient_demographics (
            max_id TEXT PRIMARY KEY,
            age INTEGER,
            gender TEXT,
            location TEXT,
            location_id TEXT
        )
    """)

    # 3. Table for patient test results
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS patient_tests (
            max_id TEXT,
            test_name TEXT,
            collection_date TEXT,
            value TEXT,
            unit TEXT,
            PRIMARY KEY (max_id, test_name, collection_date)
        )
    """)

    # 4. Table for longitudinal data
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS longitudinal_data (
            max_id TEXT,
            test_name TEXT,
            dates_values TEXT,
            PRIMARY KEY (max_id, test_name)
        )
    """)

    # Populate extractions
    for fname, meta, tdata, kft_units in payload.get('all_extractions', []):
        cursor.execute(
            "INSERT OR REPLACE INTO extractions (pdf_name, metadata, test_data, kft_units) VALUES (?, ?, ?, ?)",
            (fname, json.dumps(meta), json.dumps(tdata), json.dumps(kft_units))
        )

    # Populate demographics
    for mid, meta in payload.get('m_map', {}).items():
        cursor.execute(
            "INSERT OR REPLACE INTO patient_demographics (max_id, age, gender, location, location_id) VALUES (?, ?, ?, ?, ?)",
            (
                mid,
                meta.get('Age'),
                meta.get('Gender'),
                meta.get('Location') or meta.get('Centre'),
                meta.get('Location_ID')
            )
        )

    # Populate patient tests
    patient_data = payload.get('patient_data', {})
    kft_unit_map = payload.get('kft_unit_map', {})
    for mid, tests in patient_data.items():
        for test_name, date_vals in tests.items():
            for date, val in date_vals.items():
                unit = kft_unit_map.get(mid, {}).get(test_name, {}).get(date, 'nil')
                cursor.execute(
                    "INSERT OR REPLACE INTO patient_tests (max_id, test_name, collection_date, value, unit) VALUES (?, ?, ?, ?, ?)",
                    (mid, test_name, date, str(val), unit)
                )

    # Populate longitudinal data
    longitudinal_tests = payload.get('longitudinal_tests', {})
    for mid, test_names in longitudinal_tests.items():
        for test_name in test_names:
            dates_vals = patient_data.get(mid, {}).get(test_name, {})
            cursor.execute(
                "INSERT OR REPLACE INTO longitudinal_data (max_id, test_name, dates_values) VALUES (?, ?, ?)",
                (mid, test_name, json.dumps(dates_vals))
            )

    conn.commit()
    conn.close()


def generate_help_parameters(output_dir: Path):
    from maxscriber.constants import CBC_TESTS, LFT_TESTS, KFT_TESTS, DENGUE_TESTS
    
    help_file = output_dir / 'Max_Help_Parameters.txt'
    lines = [
        "==================================================",
        "MaxScriber Test Category Mappings",
        "==================================================",
        "",
        "CBC (Complete Blood Count):",
        *[f"  - {t}" for t in sorted(CBC_TESTS)],
        "",
        "LFT (Liver Function Test):",
        *[f"  - {t}" for t in sorted(LFT_TESTS)],
        "",
        "KFT (Kidney Function Test):",
        *[f"  - {t}" for t in sorted(KFT_TESTS)],
        "",
        "DENGUE:",
        *[f"  - {t}" for t in sorted(DENGUE_TESTS)],
        "",
        "=================================================="
    ]
    with open(help_file, 'w', encoding='utf-8') as f:
        f.write("\n".join(lines))

def run_transcribe(input_dir: Path, output_dir: Path, pattern: str = '2024_2025', threads: int = None, verbose: bool = False, job_name: str = 'default', selected_categories: list = None, progress_callback=None):
    """
    Phase 1: PDF text extraction, test mapping, and QC hashing.
    Saves intermediate data to disk for downstream subcommands.
    """
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

    output_dir.mkdir(parents=True, exist_ok=True)
    generate_help_parameters(output_dir)
    logger = setup_logging(output_dir, append=False, verbose=verbose, log_filename='extraction.log', job_name=job_name)
    logger.info(f"Input: {input_dir}")
    logger.info(f"Output: {output_dir}")

    pdf_files = sorted([f for f in input_dir.iterdir() if f.is_file() and f.suffix.lower() == '.pdf'], key=lambda x: x.name)
    if not pdf_files:
        logger.error("No PDFs found.")
        print("No PDFs found in input directory.")
        return

    all_extractions = []
    failed_files = []
    seen_hashes = {}
    qc_duplicates = []  # List of dicts: [{'duplicate': 'A.pdf', 'original': 'B.pdf'}, ...]

    logger.info(f"Found {len(pdf_files)} PDFs. Starting extraction...")
    if not verbose:
        print(f"Starting extraction for {len(pdf_files)} files (Parallel mode, worker processes)...")
    else:
        print(f"Starting extraction for {len(pdf_files)} files (Sequential verbose mode)...")
    
    start_time = time.time()

    if not verbose:
        num_workers = threads if threads is not None else os.cpu_count()
        logger.info(f"Running in parallel mode with {num_workers} processes.")
        
        with alive_bar(len(pdf_files), title='Transcribing PDFs', force_tty=True) as bar:
            with concurrent.futures.ProcessPoolExecutor(max_workers=num_workers) as executor:
                futures = {
                    executor.submit(_process_pdf_worker, pdf, output_dir, 'extraction.log', False, pattern, selected_categories): pdf
                    for pdf in pdf_files
                }
                
                for future in concurrent.futures.as_completed(futures):
                    pdf = futures[future]
                    try:
                        pdf_name, content_hash, meta, final_test_data, kft_units, confidence = future.result()
                        if progress_callback:
                            progress_callback(pdf_name, confidence)
                        
                        if content_hash is None:
                            failed_files.append(pdf_name)
                            logger.warning(f"No data extracted from {pdf_name}")
                        else:
                            # QC: Check Duplicate Content
                            if content_hash in seen_hashes:
                                orig = seen_hashes[content_hash]
                                logger.warning(f"DUPLICATE DETECTED: {pdf_name} is identical to {orig}")
                                qc_duplicates.append({'duplicate': pdf_name, 'original': orig})
                            else:
                                seen_hashes[content_hash] = pdf_name
                            
                            if final_test_data:
                                all_extractions.append((pdf_name, meta, final_test_data, kft_units))
                            else:
                                failed_files.append(pdf_name)
                                logger.warning(f"No data extracted from {pdf_name}")
                    except Exception as exc:
                        logger.error(f"Process generated an exception for {pdf.name}: {exc}", exc_info=True)
                        failed_files.append(pdf.name)
                    
                    bar()
    else:
        logger.info("Running in sequential verbose mode.")
        for idx, pdf in enumerate(pdf_files, 1):
            logger.info(f"\n{'=' * 80}")
            logger.info(f"PDF {idx}/{len(pdf_files)}: {pdf.name}")
            logger.info(f"{'=' * 80}")
            
            pdf_name, content_hash, meta, final_test_data, kft_units, confidence = _process_pdf_worker(
                pdf, output_dir, 'extraction.log', True, pattern, selected_categories
            )
            if progress_callback:
                progress_callback(pdf_name, confidence)
            
            if content_hash is None:
                failed_files.append(pdf_name)
            else:
                if content_hash in seen_hashes:
                    orig = seen_hashes[content_hash]
                    logger.warning(f"DUPLICATE DETECTED: {pdf_name} is identical to {orig}")
                    qc_duplicates.append({'duplicate': pdf_name, 'original': orig})
                else:
                    seen_hashes[content_hash] = pdf_name
                
                if final_test_data:
                    all_extractions.append((pdf_name, meta, final_test_data, kft_units))
                else:
                    failed_files.append(pdf_name)

    # Aggregate
    aggregated = _aggregate_data(all_extractions)
    
    # Identify longitudinal
    patient_data = aggregated['patient_data']
    longitudinal_tests = identify_longitudinal_tests(patient_data)

    # Save intermediate data
    payload = {
        'all_extractions': all_extractions,
        'failed_files': failed_files,
        'qc_duplicates': qc_duplicates,
        'longitudinal_tests': longitudinal_tests,
        **aggregated,
    }
    save_extraction_data(output_dir, payload, job_name=job_name)

    # Commit payload to SQLite database
    _save_to_sqlite(output_dir, payload, job_name=job_name)

    elapsed_time = time.time() - start_time
    minutes, seconds = divmod(int(elapsed_time), 60)
    
    manual_time_minutes = len(pdf_files) * 5
    
    print(f"\nTotal time taken to transcribe: {minutes} Minutes and {seconds} Seconds.")
    print(f"Doing this manually would have taken you approximately {manual_time_minutes} minutes. You're welcome! ;)")
    
    print(f"\nExtraction complete. {len(all_extractions)} files processed, {len(failed_files)} failed.")
    print(f"Check extraction.log for details.")
    logger.info("Extraction phase complete.")


# =============================================================================
# QC PHASE
# =============================================================================

def run_qc(output_dir: Path, job_name: str = 'default'):
    """
    Phase 2: Report content hashing and duplicate detection results.
    """
    data = load_extraction_data(output_dir, job_name=job_name)
    qc_duplicates = data.get('qc_duplicates', [])
    failed_files = data.get('failed_files', [])

    print("\n" + "=" * 50)
    print("QUALITY CONTROL REPORT")
    print("=" * 50)
    print(f"Duplicate Content Files : {len(qc_duplicates)}")
    if qc_duplicates:
        for item in qc_duplicates:
            print(f"    -> {item['duplicate']} (identical to {item['original']})")
    print(f"Failed/No Data Files    : {len(failed_files)}")
    if failed_files:
        for f in failed_files:
            print(f"    -> {f}")
    print("=" * 50)


# =============================================================================
# STATS PHASE
# =============================================================================

def run_stats(output_dir: Path, job_name: str = 'default'):
    """
    Phase 3: Generate the Stats_Refined.txt statistical report.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logging(output_dir, append=True, log_filename='extraction.log', job_name=job_name)

    data = load_extraction_data(output_dir, job_name=job_name)

    generate_condensed_stats(
        latest_data=data['latest_data'],
        metadata_map=data['m_map'],
        all_extractions=data['all_extractions'],
        qc_dupes=data['qc_duplicates'],
        failed_files=data['failed_files'],
        file_name_map=data['f_map'],
        longitudinal_tests=data.get('longitudinal_tests', {}),
        patient_data=data.get('patient_data', {}),
        output_path=output_dir / f'{job_name}_Stats_Refined.txt',
        logger=logger,
    )
    print(f"Stats report saved to {output_dir / f'{job_name}_Stats_Refined.txt'}")


# =============================================================================
# PLOT PHASE
# =============================================================================

def run_plot(output_dir: Path, job_name: str = 'default'):
    """
    Phase 4: Generate clinical distribution graphs.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    logger = setup_logging(output_dir, append=True, log_filename='extraction.log', job_name=job_name)

    data = load_extraction_data(output_dir, job_name=job_name)

    generate_clinical_plots(
        latest_data=data['latest_data'],
        metadata_map=data['m_map'],
        output_dir=output_dir,
        job_name=job_name,
        logger=logger,
    )
    print(f"Clinical plots saved to {output_dir / f'{job_name}_graphs'}")


# =============================================================================
# STRATIFICATION & R INTEGRATION PHASE
# =============================================================================

def run_stratification(output_dir: Path, job_name: str = 'default'):
    """
    Phase 4: Run clinical stratification and generate R analysis.
    """
    from maxscriber.stratification import load_stratification_config, apply_legacy_dengue_stratification, export_stratified_data, generate_r_script
    
    data = load_extraction_data(output_dir, job_name=job_name)
    logger = setup_logging(output_dir, append=True, log_filename='extraction.log', job_name=job_name)
    
    # Try custom yaml
    # We'll look for custom_stratification.yaml in the output_dir or current working directory
    config_path = Path.cwd() / 'custom_stratification.yaml'
    config = load_stratification_config(config_path, logger)
    
    if config:
        logger.info("Custom stratification from YAML is not fully implemented for dynamic parsing yet. Falling back to Legacy.")
        # Future: Implement dynamic yaml rule engine.
        stratified = apply_legacy_dengue_stratification(data['latest_data'], data['m_map'])
    else:
        stratified = apply_legacy_dengue_stratification(data['latest_data'], data['m_map'])
        
    xlsx_path = export_stratified_data(stratified, data['latest_data'], data['m_map'], output_dir, job_name, logger)
    generate_r_script(xlsx_path, output_dir, job_name, logger)
    
    print(f"Stratification data saved to {xlsx_path.name}")
    print(f"R script generated at {job_name}_analysis.R")


# =============================================================================
# FULL PIPELINE
# =============================================================================

def run_all_phase1(input_dir: Path, output_dir: Path, pattern: str = '2024_2025', threads: int = None, verbose: bool = False, job_name: str = 'default', selected_categories: list = None):
    """
    Executes the entire Phase 1 pipeline:
      - PDF extraction & text parsing
      - QC Deduplication
      - JSON intermediate save
      - Excel Master_Data_Refined.xlsx generation
      - Longitudinal data extraction
    """
    start = time.time()
    logger = logging.getLogger('MaxScriber')
    
    run_transcribe(input_dir, output_dir, pattern=pattern, threads=threads, verbose=verbose, job_name=job_name, selected_categories=selected_categories)

    # Load aggregated data and generate Excel report
    # Load aggregated data
    data = load_extraction_data(output_dir, job_name=job_name)
    logger = setup_logging(output_dir, append=True, log_filename='extraction.log', job_name=job_name)
    
    # Generate Master Excel
    _generate_excel_report(data, output_dir, logger, pattern=pattern, job_name=job_name)
    
    # Generate Longitudinal Excel if data exists
    longitudinal_tests = data.get('longitudinal_tests', {})
    if longitudinal_tests:
        generate_longitudinal_excel(
            data['patient_data'], 
            longitudinal_tests, 
            output_dir / f'{job_name}_longitudinal_data.xlsx', 
            logger
        )

    # Generate Lab Variance Report
    generate_lab_variance_report(
        centre_map=data.get('centre_map', {}),
        patient_data=data.get('patient_data', {}),
        f_map=data.get('f_map', {}),
        kft_unit_map=data.get('kft_unit_map', {}),
        output_path=output_dir / f'{job_name}_lab_variance_report.txt',
        logger=logger,
    )
    
    logger.info("Phase 1 complete: Transcription and Excel reports generated.")


def run_all_phase2(output_dir: Path, job_name: str = 'default'):
    """
    Phase 2 of the full pipeline: QC + Stats + Plot.
    Called after the verification gate approves continuation.
    """
    data = load_extraction_data(output_dir, job_name=job_name)
    logger = setup_logging(output_dir, append=True, log_filename='extraction.log', job_name=job_name)

    # QC (terminal output)
    run_qc(output_dir, job_name=job_name)

    # Stats (pass longitudinal info)
    generate_condensed_stats(
        latest_data=data['latest_data'],
        metadata_map=data['m_map'],
        all_extractions=data['all_extractions'],
        qc_dupes=data['qc_duplicates'],
        failed_files=data['failed_files'],
        file_name_map=data['f_map'],
        longitudinal_tests=data.get('longitudinal_tests', {}),
        patient_data=data.get('patient_data', {}),
        output_path=output_dir / f'{job_name}_Stats_Refined.txt',
        logger=logger,
    )

    generate_clinical_plots(
        latest_data=data['latest_data'],
        metadata_map=data['m_map'],
        output_dir=output_dir,
        job_name=job_name,
        logger=logger,
    )

    # Stratification & R Analysis
    run_stratification(output_dir, job_name=job_name)

    logger.info("DONE.")
    print(f"\nAll reports saved to {output_dir}")


# =============================================================================
# HELPERS
# =============================================================================

def _aggregate_data(all_extractions: List) -> dict:
    """Aggregate per-PDF extractions into per-patient data."""
    p_data = defaultdict(lambda: defaultdict(dict))
    f_map = defaultdict(set)
    m_map = {}
    # centre_map: MAX_id -> {date -> {centre, fname}} for variance detection
    centre_map: Dict[str, List[dict]] = defaultdict(list)
    # kft_unit_map: MAX_id -> test -> {date -> unit}
    kft_unit_map: Dict[str, Dict[str, Dict[str, str]]] = defaultdict(lambda: defaultdict(dict))

    for record in all_extractions:
        # Support both old 3-tuple (backward compat) and new 4-tuple
        if len(record) == 4:
            fname, meta, tdata, kft_units = record
        else:
            fname, meta, tdata = record
            kft_units = {}

        mid = meta.get('MAX_id') or f"UNKNOWN_{fname}"
        f_map[mid].add(fname)
        if mid not in m_map or (not m_map[mid].get('Age') and meta.get('Age')):
            m_map[mid] = meta

        # Track centre per date for variance reporting
        centre = meta.get('Centre')
        for t, dv in tdata.items():
            for d, v in dv.items():
                p_data[mid][t][d] = v
                if centre:
                    centre_map[mid].append({
                        'date': d,
                        'centre': centre,
                        'fname': fname,
                    })

        # Merge KFT units
        for test_nm, date_unit in kft_units.items():
            for d, u in date_unit.items():
                kft_unit_map[mid][test_nm][d] = u

    latest_data = {}
    for mid, tests in p_data.items():
        latest_data[mid] = {}
        last_dt = None
        for t, dv in tests.items():
            valid = []
            for dstr, val in dv.items():
                try:
                    dt = datetime.strptime(dstr, '%d-%m-%Y')
                    valid.append((dt, dstr, val))
                except Exception:
                    pass
            if valid:
                valid.sort(reverse=True)
                latest_data[mid][t] = valid[0][2]
                if not last_dt or valid[0][0] > last_dt:
                    last_dt = valid[0][0]
                    latest_data[mid]['collection_date'] = valid[0][1]
            else:
                latest_data[mid][t] = 'nil'

    return {
        'patient_data': dict(p_data),
        'latest_data': latest_data,
        'f_map': dict(f_map),
        'm_map': m_map,
        'centre_map': dict(centre_map),
        'kft_unit_map': dict(kft_unit_map),
    }


def identify_longitudinal_tests(patient_data: Dict) -> Dict:
    """
    Identify tests with MULTIPLE dates (>= 2 distinct dates).
    Refined Definition: Any MaxID that has at least one identical clinical test 
    recorded on two or more different Collection Dates.
    """
    longitudinal = {}

    for max_id, tests in patient_data.items():
        multi_date = []
        for test_name, date_values in tests.items():
            # Get valid distinct dates
            raw_dates = set([d for d in date_values.keys() if d != 'nil'])
            
            # Simple check: Are there 2 or more distinct dates?
            if len(raw_dates) >= 2:
                multi_date.append(test_name)

        if multi_date:
            longitudinal[max_id] = multi_date

    return longitudinal


def _generate_excel_report(data: dict, output_dir: Path, logger: logging.Logger, pattern: str = '2024_2025', job_name: str = 'default'):
    """Generate Master_Data_Refined.xlsx with branding and highlighting."""
    latest_data = data['latest_data']
    f_map = data['f_map']
    m_map = data['m_map']
    longitudinal_tests = data.get('longitudinal_tests', {})

    rows = []
    # Sort by MAX_id for consistency
    sorted_mids = sorted(latest_data.keys())
    
    from maxscriber.constants import TEST_ALIASES

    # Collect data for all rows
    for mid in sorted_mids:
        row = {
            'MAX_id': mid,
            'File Name': ', '.join(sorted(list(f_map[mid]))),
            'Tests Done': determine_tests_done(latest_data[mid]),
            'SIN No.': m_map.get(mid, {}).get('SIN_No', 'nil'),
            'Gender': m_map.get(mid, {}).get('Gender', 'nil'),
            'Age': m_map.get(mid, {}).get('Age', 'nil'),
            'Collection_date': latest_data[mid].get('collection_date', 'nil'),
        }

        if pattern == '2023':
            row['Location ID'] = m_map.get(mid, {}).get('Location_ID', '-')
            row['Location '] = m_map.get(mid, {}).get('Location', '-')
        else:
            row['Centre'] = m_map.get(mid, {}).get('Centre', 'nil')

        for c in list(TEST_ALIASES.keys()):
            row[c] = latest_data[mid].get(c, 'nil')
            # Extract corresponding unit if tracking
            col_date = latest_data[mid].get('collection_date')
            unit_map = data.get('kft_unit_map', {})
            unit_val = unit_map.get(mid, {}).get(c, {}).get(col_date, 'nil') if col_date else 'nil'
            row[f"{c}_Unit"] = unit_val
        rows.append(row)

    df = pd.DataFrame(rows)

    if pattern == '2023':
        template_columns = [
            'MAX_id', 'Location ID', 'Location ', 'File Name', 'Tests Done', 'SIN No.', 'Gender', 
            'Age', 'Collection_date', 'Haemoglobin', 'Packed Cell,Volume', 'Total Leukocyte Count(TLC)', 
            'RBC', 'MCV', 'MCH', 'MCHC', 'Platelet', 'MPV', 'RDW', 'Neutrophils', 'Lymphocytes', 
            'Monocytes', 'Eosinophils', 'Basophils', 'Absolute Neutrophil Count', 'Absolute Lymphocyte Count', 
            'Absolute Monocyte Count', 'Absolute Eosinophil Count', 'Absolute Basophil Count', 
            'Total Protein', 'Albumin', 'Globulin', 'A.G.', 'Bilirubin (Total)', 'Bilirubin (Direct)', 
            'Bilirubin (Indirect)', 'Transaminase (AST)', 'Transaminase (ALT)', 'Alkaline Phosphatase', 
            'GGTP (Gamma GT), Serum', 'Dengue NS1 Antigen', 'Dengue IgG', 'Dengue IgM', 'Urea', 'BUN', 
            'Creatinine', 'BUN_Creatinine_Ratio', 'Uric_Acid', 'eGFR'
        ]
        
        # Determine which test columns actually have non-'nil' data
        metadata_cols = ['MAX_id', 'Location ID', 'Location ', 'File Name', 'Tests Done', 'SIN No.', 'Gender', 'Age', 'Collection_date']
        test_cols_in_template = [c for c in template_columns if c not in metadata_cols]
        
        valid_test_cols = []
        for col in test_cols_in_template:
            if col in df.columns:
                # Check if there is any non-'nil' value in this column
                if not (df[col] == 'nil').all():
                    valid_test_cols.append(col)
                    
        # Only include test columns from the template (no Unit columns)
        final_cols = metadata_cols + valid_test_cols
        df = df[final_cols]
    else:
        # Legacy pattern: Exclude Unit columns from final output
        legacy_cols = [c for c in OUTPUT_COLUMNS if not c.endswith('_Unit')]
        df = df[legacy_cols]
    xlsx_path = output_dir / f'{job_name}_Master_Data_Refined.xlsx'
    
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Master')
        ws = writer.sheets['Master']
        
        # Styles
        header_font = Font(bold=True)
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Bold Headers
        for cell in ws[1]:
            cell.font = header_font

        # Highlight Longitudinal MAX_id (Column A)
        # rows are 1-indexed. Header is row 1. Data starts row 2.
        for row_idx, mid in enumerate(df['MAX_id'], start=2):
            if mid in longitudinal_tests:
                ws.cell(row=row_idx, column=1).fill = yellow_fill

        # Highlight Dengue Positives
        dengue_thresholds = {
            'Dengue NS1 Antigen': 0.9,
            'Dengue IgG': 9.0,
            'Dengue IgM': 9.0,
        }
        
        # Map column names to indices (1-based)
        col_map = {name: idx for idx, name in enumerate(OUTPUT_COLUMNS, start=1)}
        
        for col_name, threshold in dengue_thresholds.items():
            if col_name in col_map:
                col_idx = col_map[col_name]
                for row_idx in range(2, len(df) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val and val != 'nil' and str(val).strip():
                        try:
                            # Clean string value to float
                            clean_val = str(val).lower().replace('>', '').replace('<', '').strip()
                            if clean_val in ['positive', 'reactive', 'detected']:
                                cell.fill = yellow_fill
                            else:
                                fval = float(clean_val)
                                if fval >= threshold:
                                    cell.fill = yellow_fill
                        except ValueError:
                            pass

    logger.info(f"Excel report saved: {xlsx_path}")


def generate_longitudinal_excel(patient_data: Dict, longitudinal_tests: Dict, output_path: Path, logger: logging.Logger):
    """Generate longitudinal_data.xlsx for patients with multi-date records."""
    rows = []

    for max_id, test_names in longitudinal_tests.items():
        for test_name in test_names:
            date_values = patient_data[max_id][test_name]

            sorted_items = []
            for date_str, value in date_values.items():
                try:
                    dt = datetime.strptime(date_str, '%d-%m-%Y')
                    sorted_items.append((dt, date_str, value))
                except ValueError:
                    pass

            sorted_items.sort() # sort by date ascending

            for _, date_str, value in sorted_items:
                rows.append({
                    'MAX_id': max_id,
                    'Test_Name': test_name,
                    'Collection_Date': date_str,
                    'Value': value
                })

    if not rows:
        return

    df = pd.DataFrame(rows, columns=['MAX_id', 'Test_Name', 'Collection_Date', 'Value'])
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Longitudinal')
        ws = writer.sheets['Longitudinal']
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
    logger.info(f"Longitudinal report saved: {output_path} ({len(rows)} records)")

# =============================================================================
# LAB VARIANCE REPORT
# =============================================================================

def _normalize_centre_name(centre: str) -> str:
    if not centre:
        return ""
    import re
    s = centre.lower().strip()
    # Strip numeric prefix
    s = re.sub(r'^\d+\s*-\s*', '', s)
    # Remove parenthesized details
    s = re.sub(r'\(.*?\)', '', s)
    # Remove noise suffixes
    s = re.sub(r'\b(company|owned?|own|cent(?:er|re)|lab|labs)\b', '', s)
    # Strip non-alphanumeric chars
    s = re.sub(r'[^a-z0-9]', '', s)
    return s.strip()


def generate_lab_variance_report(
    centre_map: Dict,
    patient_data: Dict,
    f_map: Dict,
    kft_unit_map: Dict,
    output_path: Path,
    logger: logging.Logger,
):
    """
    Detect patients whose KFT results were obtained from multiple lab centres,
    then emit a structured lab_variance_report.txt.

    A Lab-Based Change is flagged when the same MAX_ID has entries linked to
    two or more different Centre values, indicating:
      - A shift in testing centre between visits, OR
      - Differing lab methodology or units for the same test.
    """
    from maxscriber.constants import KFT_TESTS
    from datetime import datetime as _dt

    flagged = []

    for max_id, entries in centre_map.items():
        # Deduplicate by (date, centre, fname)
        seen = set()
        unique_entries = []
        for e in entries:
            key = (e["date"], e["centre"], e["fname"])
            if key not in seen:
                seen.add(key)
                unique_entries.append(e)

        # Distinct centres for this patient using normalized comparison
        normalized_centres = {_normalize_centre_name(e["centre"]) for e in unique_entries}
        if len(normalized_centres) < 2:
            continue  # Same physical centre for all visits

        detail_lines = []
        tests_examined = patient_data.get(max_id, {})

        for test_name in sorted(KFT_TESTS):
            if test_name not in tests_examined:
                continue
            date_values = tests_examined[test_name]
            for date_str, value in sorted(date_values.items()):
                matching = [e for e in unique_entries if e["date"] == date_str]
                centre_label = matching[0]["centre"] if matching else "Unknown"
                fname_label = matching[0]["fname"] if matching else "Unknown"
                unit = (kft_unit_map
                        .get(max_id, {})
                        .get(test_name, {})
                        .get(date_str, ""))
                result_str = (f"{value} {unit}").strip()
                detail_lines.append(
                    f"  Date: {date_str} | Centre: {centre_label} | "
                    f"Test: {test_name} | Result: {result_str} | File: {fname_label}"
                )

        # Fallback if no KFT tests found
        if not detail_lines:
            for e in unique_entries:
                detail_lines.append(
                    "  Date: {} | Centre: {} | Test: (Non-KFT) | Result: N/A | File: {}".format(
                        e["date"], e["centre"], e["fname"]
                    )
                )

        # Determine observation type
        obs = "Shift in Testing Centre"
        for test_name in KFT_TESTS:
            units_seen = set(
                kft_unit_map.get(max_id, {}).get(test_name, {}).values()
            )
            if len(units_seen) > 1:
                obs = "Change in Methodology / Shift in Testing Centre"
                break

        source_pdfs = sorted(f_map.get(max_id, set()))
        flagged.append({
            "max_id": max_id,
            "observation": obs,
            "detail_lines": detail_lines,
            "source_pdfs": source_pdfs,
        })

    # Write report file
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("=" * 70 + "\n")
        f.write("LAB VARIANCE & OBSERVATION REPORT\n")
        f.write("Generated: {}\n".format(_dt.now().strftime("%Y-%m-%d %H:%M:%S")))
        f.write("Total Flagged Patients: {}\n".format(len(flagged)))
        f.write("=" * 70 + "\n\n")

        if not flagged:
            f.write("No lab-based variances detected.\n")
        else:
            for entry in flagged:
                f.write("Patient ID  : {}\n".format(entry["max_id"]))
                f.write("Observation : {}\n".format(entry["observation"]))
                f.write("Details:\n")
                for line in entry["detail_lines"]:
                    f.write(line + "\n")
                f.write(
                    "Tracing     : Source PDFs: "
                    + ", ".join(entry["source_pdfs"]) + "\n"
                )
                f.write("-" * 70 + "\n\n")

    logger.info(
        "Lab variance report saved: {} ({} flagged patients)".format(
            output_path, len(flagged)
        )
    )
    print("Lab variance report saved to {} ({} flagged)".format(
        output_path, len(flagged)
    ))
