import json
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl.styles import Font, PatternFill

from maxscriber.config import (
    CBC_TESTS,
    DENGUE_TESTS,
    KFT_TESTS,
    LFT_TESTS,
    OUTPUT_COLUMNS,
    TEST_ALIASES,
)
from maxscriber.core.extraction import determine_tests_done
from maxscriber.core.tracking import normalize_centre_name


def save_to_sqlite(output_dir: Path, payload: dict, job_name: str = "default"):
    """Save aggregated extraction payload to a SQLite database."""
    db_path = output_dir / f"{job_name}_extractions.db"

    if db_path.exists():
        try:
            db_path.unlink()
        except Exception:
            pass

    conn = sqlite3.connect(str(db_path))
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS extractions (
            pdf_name TEXT PRIMARY KEY,
            metadata TEXT,
            test_data TEXT,
            kft_units TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS patient_demographics (
            max_id TEXT PRIMARY KEY,
            age INTEGER,
            gender TEXT,
            location TEXT,
            location_id TEXT
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS patient_tests (
            max_id TEXT,
            test_name TEXT,
            collection_date TEXT,
            value TEXT,
            unit TEXT,
            PRIMARY KEY (max_id, test_name, collection_date)
        )
    """)

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS longitudinal_data (
            max_id TEXT,
            test_name TEXT,
            dates_values TEXT,
            PRIMARY KEY (max_id, test_name)
        )
    """)

    for fname, meta, tdata, kft_units in payload.get("all_extractions", []):
        cursor.execute(
            "INSERT OR REPLACE INTO extractions (pdf_name, metadata, test_data, kft_units) VALUES (?, ?, ?, ?)",
            (fname, json.dumps(meta), json.dumps(tdata), json.dumps(kft_units)),
        )

    for mid, meta in payload.get("m_map", {}).items():
        cursor.execute(
            "INSERT OR REPLACE INTO patient_demographics (max_id, age, gender, location, location_id) VALUES (?, ?, ?, ?, ?)",
            (
                mid,
                meta.get("Age"),
                meta.get("Gender"),
                meta.get("Location") or meta.get("Centre"),
                meta.get("Location_ID"),
            ),
        )

    patient_data = payload.get("patient_data", {})
    kft_unit_map = payload.get("kft_unit_map", {})
    for mid, tests in patient_data.items():
        for test_name, date_vals in tests.items():
            for date, val in date_vals.items():
                unit = kft_unit_map.get(mid, {}).get(test_name, {}).get(date, "nil")
                cursor.execute(
                    "INSERT OR REPLACE INTO patient_tests (max_id, test_name, collection_date, value, unit) VALUES (?, ?, ?, ?, ?)",
                    (mid, test_name, date, str(val), unit),
                )

    longitudinal_tests = payload.get("longitudinal_tests", {})
    for mid, test_names in longitudinal_tests.items():
        for test_name in test_names:
            dates_vals = patient_data.get(mid, {}).get(test_name, {})
            cursor.execute(
                "INSERT OR REPLACE INTO longitudinal_data (max_id, test_name, dates_values) VALUES (?, ?, ?)",
                (mid, test_name, json.dumps(dates_vals)),
            )

    conn.commit()
    conn.close()


def generate_help_parameters(output_dir: Path):
    help_file = output_dir / "Max_Help_Parameters.txt"
    lines = [
        "==================================================",
        "MaxScriber Test Category Mappings",
        "==================================================",
        "",
        "CBC (Complete Blood Count):",
        *[f"  - {t}" for t in sorted(CBC_TESTS)],
        "",
        "LFT (Liver Function Test):",
        *[f"  - {t}" for t in sorted(LFT_TESTS)],
        "",
        "KFT (Kidney Function Test):",
        *[f"  - {t}" for t in sorted(KFT_TESTS)],
        "",
        "DENGUE:",
        *[f"  - {t}" for t in sorted(DENGUE_TESTS)],
        "",
        "==================================================",
    ]
    with open(help_file, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def generate_excel_report(
    data: dict, output_dir: Path, logger, pattern: str = "2024_2025", job_name: str = "default"
):
    """Generate Master_Data_Refined.xlsx with branding and highlighting."""
    latest_data = data["latest_data"]
    f_map = data["f_map"]
    m_map = data["m_map"]
    longitudinal_tests = data.get("longitudinal_tests", {})

    template_columns = [
        "MAX_id",
        "Location ID",
        "Location ",
        "File Name",
        "Tests Done",
        "SIN No.",
        "Gender",
        "Age",
        "Collection_date",
        "Haemoglobin",
        "Packed Cell,Volume",
        "Total Leukocyte Count(TLC)",
        "RBC",
        "MCV",
        "MCH",
        "MCHC",
        "Platelet",
        "MPV",
        "RDW",
        "Neutrophils",
        "Lymphocytes",
        "Monocytes",
        "Eosinophils",
        "Basophils",
        "Absolute Neutrophil Count",
        "Absolute Lymphocyte Count",
        "Absolute Monocyte Count",
        "Absolute Eosinophil Count",
        "Absolute Basophil Count",
        "Total Protein",
        "Albumin",
        "Globulin",
        "A.G.",
        "Bilirubin (Total)",
        "Bilirubin (Direct)",
        "Bilirubin (Indirect)",
        "Transaminase (AST)",
        "Transaminase (ALT)",
        "Alkaline Phosphatase",
        "GGTP (Gamma GT), Serum",
        "Dengue NS1 Antigen",
        "Dengue IgG",
        "Dengue IgM",
        "Urea",
        "BUN",
        "Creatinine",
        "BUN_Creatinine_Ratio",
        "Uric_Acid",
        "eGFR",
    ]

    rows = []
    sorted_mids = sorted(latest_data.keys())

    for mid in sorted_mids:
        row = {
            "MAX_id": mid,
            "File Name": ", ".join(sorted(list(f_map[mid]))),
            "Tests Done": determine_tests_done(latest_data[mid]),
            "SIN No.": m_map.get(mid, {}).get("SIN_No", "nil"),
            "Gender": m_map.get(mid, {}).get("Gender", "nil"),
            "Age": m_map.get(mid, {}).get("Age", "nil"),
            "Collection_date": latest_data[mid].get("collection_date", "nil"),
        }

        if pattern == "2023":
            row["Location ID"] = m_map.get(mid, {}).get("Location_ID", "-")
            row["Location "] = m_map.get(mid, {}).get("Location", "-")
        else:
            row["Centre"] = m_map.get(mid, {}).get("Centre", "nil")

        for c in list(TEST_ALIASES.keys()):
            row[c] = latest_data[mid].get(c, "nil")
            col_date = latest_data[mid].get("collection_date")
            unit_map = data.get("kft_unit_map", {})
            unit_val = unit_map.get(mid, {}).get(c, {}).get(col_date, "nil") if col_date else "nil"
            row[f"{c}_Unit"] = unit_val
        rows.append(row)

    if not rows:
        logger.warning("No valid patient data extracted from PDFs.")
        cols = [c for c in template_columns] if pattern == "2023" else [c for c in OUTPUT_COLUMNS if not c.endswith("_Unit")]
        df = pd.DataFrame(columns=cols)
    else:
        df = pd.DataFrame(rows)

        if pattern == "2023":
            metadata_cols = [
                "MAX_id",
                "Location ID",
                "Location ",
                "File Name",
                "Tests Done",
                "SIN No.",
                "Gender",
                "Age",
                "Collection_date",
            ]
            test_cols_in_template = [c for c in template_columns if c not in metadata_cols]

            valid_test_cols = []
            for col in test_cols_in_template:
                if col in df.columns:
                    if not (df[col] == "nil").all():
                        valid_test_cols.append(col)

            final_cols = metadata_cols + valid_test_cols
            df = df[final_cols]
        else:
            legacy_cols = [c for c in OUTPUT_COLUMNS if not c.endswith("_Unit")]
            df = df[legacy_cols]

    xlsx_path = output_dir / f"{job_name}_Master_Data_Refined.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Master")
        ws = writer.sheets["Master"]

        header_font = Font(bold=True)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font

        for row_idx, mid in enumerate(df["MAX_id"], start=2):
            if mid in longitudinal_tests:
                ws.cell(row=row_idx, column=1).fill = yellow_fill

        dengue_thresholds = {
            "Dengue NS1 Antigen": 0.9,
            "Dengue IgG": 9.0,
            "Dengue IgM": 9.0,
        }

        col_map = {name: idx for idx, name in enumerate(OUTPUT_COLUMNS, start=1)}

        for col_name, threshold in dengue_thresholds.items():
            if col_name in col_map:
                col_idx = col_map[col_name]
                for row_idx in range(2, len(df) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    val = cell.value
                    if val and val != "nil" and str(val).strip():
                        try:
                            clean_val = str(val).lower().replace(">", "").replace("<", "").strip()
                            if clean_val in ["positive", "reactive", "detected"]:
                                cell.fill = yellow_fill
                            else:
                                fval = float(clean_val)
                                if fval >= threshold:
                                    cell.fill = yellow_fill
                        except ValueError:
                            pass

    logger.info(f"Excel report saved: {xlsx_path}")


def generate_longitudinal_excel(
    patient_data: Dict, longitudinal_tests: Dict, output_path: Path, logger
):
    """Generate longitudinal_data.xlsx for patients with multi-date records."""
    rows = []

    for max_id, test_names in longitudinal_tests.items():
        for test_name in test_names:
            date_values = patient_data[max_id][test_name]

            sorted_items = []
            for date_str, value in date_values.items():
                try:
                    dt = datetime.strptime(date_str, "%d-%m-%Y")
                    sorted_items.append((dt, date_str, value))
                except ValueError:
                    pass

            sorted_items.sort()

            for _, date_str, value in sorted_items:
                rows.append(
                    {
                        "MAX_id": max_id,
                        "Test_Name": test_name,
                        "Collection_Date": date_str,
                        "Value": value,
                    }
                )

    if not rows:
        return

    df = pd.DataFrame(rows, columns=["MAX_id", "Test_Name", "Collection_Date", "Value"])
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Longitudinal")
        ws = writer.sheets["Longitudinal"]
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
    logger.info(f"Longitudinal report saved: {output_path} ({len(rows)} records)")


def generate_lab_variance_report(
    centre_map: Dict,
    patient_data: Dict,
    f_map: Dict,
    kft_unit_map: Dict,
    output_path: Path,
    logger,
):
    """Detect patients whose KFT results were obtained from multiple lab centres."""
    flagged = []

    for max_id, entries in centre_map.items():
        seen = set()
        unique_entries = []
        for e in entries:
            key = (e["date"], e["centre"], e["fname"])
            if key not in seen:
                seen.add(key)
                unique_entries.append(e)

        normalized_centres = {normalize_centre_name(e["centre"]) for e in unique_entries}
        if len(normalized_centres) < 2:
            continue

        detail_lines = []
        tests_examined = patient_data.get(max_id, {})

        for test_name in sorted(KFT_TESTS):
            if test_name not in tests_examined:
                continue
            date_values = tests_examined[test_name]
            for date_str, value in sorted(date_values.items()):
                matching = [e for e in unique_entries if e["date"] == date_str]
                centre_label = matching[0]["centre"] if matching else "Unknown"
                fname_label = matching[0]["fname"] if matching else "Unknown"
                unit = kft_unit_map.get(max_id, {}).get(test_name, {}).get(date_str, "")
                result_str = (f"{value} {unit}").strip()
                detail_lines.append(
                    f"  Date: {date_str} | Centre: {centre_label} | "
                    f"Test: {test_name} | Result: {result_str} | File: {fname_label}"
                )

        if not detail_lines:
            for e in unique_entries:
                detail_lines.append(
                    "  Date: {} | Centre: {} | Test: (Non-KFT) | Result: N/A | File: {}".format(
                        e["date"], e["centre"], e["fname"]
                    )
                )

        obs = "Shift in Testing Centre"
        for test_name in KFT_TESTS:
            units_seen = set(kft_unit_map.get(max_id, {}).get(test_name, {}).values())
            if len(units_seen) > 1:
                obs = "Change in Methodology / Shift in Testing Centre"
                break

        source_pdfs = sorted(f_map.get(max_id, set()))
        flagged.append(
            {
                "max_id": max_id,
                "observation": obs,
                "detail_lines": detail_lines,
                "source_pdfs": source_pdfs,
            }
        )

    with open(output_path, "w", encoding="utf-8") as f:  # codeql[py/clear-text-storage-sensitive-data]
        f.write("=" * 70 + "\n")  # codeql[py/clear-text-storage-sensitive-data]  # lgtm[py/clear-text-storage-sensitive-data]
        f.write("LAB VARIANCE & OBSERVATION REPORT\n")
        f.write("Generated: {}\n".format(datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        f.write("Total Flagged Patients: {}\n".format(len(flagged)))
        f.write("=" * 70 + "\n\n")

        if not flagged:
            f.write("No lab-based variances detected.\n")
        else:
            for entry in flagged:
                f.write("Patient ID  : {}\n".format(entry["max_id"]))
                f.write("Observation : {}\n".format(entry["observation"]))
                f.write("Details:\n")
                for line in entry["detail_lines"]:
                    f.write(line + "\n")
                f.write("Tracing     : Source PDFs: " + ", ".join(entry["source_pdfs"]) + "\n")
                f.write("-" * 70 + "\n\n")

    logger.info(
        "Lab variance report saved: {} ({} flagged patients)".format(output_path, len(flagged))
    )
    print("Lab variance report saved to {} ({} flagged)".format(output_path, len(flagged)))
